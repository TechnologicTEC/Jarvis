"""Read and update the internship tracker spreadsheet (openpyxl).

Writing is deliberately conservative — this is the user's real, hand-maintained
file, not something Jarvis owns:

  * Never called automatically. The spec's rule is that a detected reply is
    surfaced for confirmation and only written after an explicit "log it".
  * A timestamped backup is taken before the first write of each session.
  * Only the Stage/Status cells of an already-existing company row are touched.
    Jarvis never inserts, reorders, deletes, or rewrites other columns.
  * If the workbook is open in Excel the save fails cleanly with a message
    telling the user to close it, rather than half-writing.
"""
import datetime
import os
import re
import shutil

from core import config

_backed_up = False

# Header names we understand, lowercased. The sheet is the user's, so match
# loosely rather than demanding exact spelling.
_COL_ALIASES = {
    "company": ("company", "employer", "organisation", "organization"),
    "stage": ("stage",),
    "status": ("status",),
    "date": ("dated submitted", "date submitted", "date", "submitted"),
    "location": ("location",),
}


def tracker_path() -> str:
    return os.path.expandvars(config.get("inbox", "tracker_path", default="") or "")


def is_available() -> bool:
    p = tracker_path()
    return bool(p) and os.path.isfile(p)


def _columns(header_row) -> dict:
    """Map our logical names to 1-based column indexes."""
    cols = {}
    for idx, cell in enumerate(header_row, start=1):
        name = str(cell or "").strip().lower()
        if not name:
            continue
        for key, aliases in _COL_ALIASES.items():
            if key not in cols and name in aliases:
                cols[key] = idx
    return cols


def read_applications() -> dict:
    """Every row of the tracker, as dicts. Read-only."""
    path = tracker_path()
    if not path or not os.path.isfile(path):
        return {"ok": False, "error": "no_tracker", "applications": []}
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        return {"ok": False, "error": str(e), "applications": []}
    try:
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return {"ok": True, "applications": [], "columns": {}}
        cols = _columns(rows[0])
        if "company" not in cols:
            return {"ok": False, "error": "no_company_column", "applications": []}
        out = []
        for i, row in enumerate(rows[1:], start=2):
            company = row[cols["company"] - 1] if len(row) >= cols["company"] else None
            if not company or not str(company).strip():
                continue

            def val(key):
                idx = cols.get(key)
                if not idx or len(row) < idx:
                    return ""
                v = row[idx - 1]
                if isinstance(v, (datetime.datetime, datetime.date)):
                    return v.strftime("%Y-%m-%d")
                return "" if v is None else str(v).strip()

            out.append({
                "row": i, "company": str(company).strip(),
                "stage": val("stage"), "status": val("status"),
                "date": val("date"), "location": val("location"),
            })
        return {"ok": True, "applications": out, "columns": cols}
    finally:
        wb.close()


def companies() -> list:
    r = read_applications()
    return [a["company"] for a in r.get("applications", [])]


def _norm(s: str) -> str:
    return re.sub(r"[^a-z0-9]", "", (s or "").lower())


def find_company(name: str):
    """Match a company name from an email to a tracker row, loosely."""
    target = _norm(name)
    if not target:
        return None
    apps = read_applications().get("applications", [])
    for a in apps:
        if _norm(a["company"]) == target:
            return a
    for a in apps:
        c = _norm(a["company"])
        if c and (c in target or target in c):
            return a
    return None


def _backup(path: str):
    global _backed_up
    if _backed_up:
        return
    stamp = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
    base, ext = os.path.splitext(path)
    try:
        shutil.copy2(path, f"{base}.jarvis-backup-{stamp}{ext}")
        _backed_up = True
    except Exception:
        pass  # a failed backup must not block the user's explicit request


def update_stage(company: str, stage: str = None, status: str = None) -> dict:
    """Set Stage/Status on an existing company row. Explicit action only."""
    path = tracker_path()
    if not path or not os.path.isfile(path):
        return {"ok": False, "reply": "No tracker configured — set inbox.tracker_path"}

    hit = find_company(company)
    if not hit:
        known = ", ".join(companies()[:6]) or "none"
        return {"ok": False,
                "reply": f"No row for “{company}” in the tracker (have: {known})"}

    try:
        import openpyxl
        wb = openpyxl.load_workbook(path)
    except Exception as e:
        return {"ok": False, "reply": f"Could not open the tracker — {e}"}
    try:
        ws = wb[wb.sheetnames[0]]
        cols = _columns(next(ws.iter_rows(max_row=1, values_only=True)))
        changed = []
        if stage and cols.get("stage"):
            ws.cell(row=hit["row"], column=cols["stage"]).value = stage
            changed.append(f"stage → {stage}")
        if status and cols.get("status"):
            ws.cell(row=hit["row"], column=cols["status"]).value = status
            changed.append(f"status → {status}")
        if not changed:
            return {"ok": False, "reply": "Nothing to change (no Stage/Status column found)"}
        _backup(path)
        wb.save(path)
    except PermissionError:
        return {"ok": False,
                "reply": "The tracker is open in Excel — close it and try again"}
    except Exception as e:
        return {"ok": False, "reply": f"Could not write the tracker — {e}"}
    finally:
        wb.close()

    return {"ok": True, "company": hit["company"], "row": hit["row"],
            "reply": f"Logged {hit['company']}: " + ", ".join(changed)}


def summary() -> dict:
    r = read_applications()
    if not r.get("ok"):
        err = r.get("error")
        if err == "no_tracker":
            return {"ok": False, "intent": "mail",
                    "reply": "No tracker configured — set inbox.tracker_path in settings.json"}
        return {"ok": False, "intent": "mail", "reply": f"Tracker unreadable — {err}"}
    apps = r["applications"]
    if not apps:
        return {"ok": True, "intent": "mail", "reply": "Tracker is empty.", "applications": []}
    pending = [a for a in apps if not a["status"] or a["status"].strip("_ ") == ""]
    return {"ok": True, "intent": "mail",
            "reply": f"{len(apps)} application(s), {len(pending)} awaiting a reply — "
                     + ", ".join(a["company"] for a in apps[:4]),
            "applications": apps}
