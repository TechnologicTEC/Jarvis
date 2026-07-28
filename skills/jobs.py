"""Finding internships worth applying to — and hiding the ones you can't.

The filter is the point. Most "software internship Auckland" results are for
penultimate or final-year students, and scrolling past those is the actual
chore. You're 2nd year of 4, so anything demanding 3rd/4th year, a specific
graduation year, or "penultimate" is dropped before you ever see it.

Discovery goes through Tavily rather than scraping job boards: it respects
robots, returns extracted page content, and doesn't break when a site changes
its markup. Seek's own JSON search endpoint was tried first and 404s now,
which is exactly the fragility worth avoiding.

Results are cached, because each search costs API credits and job boards do
not change minute to minute.
"""
import os
import re
import threading
import time
import urllib.parse

from core import config

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# Where NZ student tech roles actually get posted.
DEFAULT_SITES = [
    "seek.co.nz", "nz.indeed.com", "nz.linkedin.com", "sjs.co.nz",
    "summeroftech.co.nz", "nz.gradconnection.com", "trademe.co.nz",
    "workhere.co.nz", "joblist.co.nz",
]

# Role wording to search for. Kept broad across engineering/CS but always
# anchored to software or computing, per what you asked for.
def _season_terms():
    """The season that's actually open for applications.

    NZ summer internships run November-February and recruit through the middle
    of the year, so from about March the live season is the *next* one. Without
    this the searches kept surfacing the season just gone.
    """
    import datetime
    t = datetime.date.today()
    start = t.year if t.month >= 3 else t.year - 1
    return f"{start}/{start + 1}", f"{start} {start + 1}", str(start + 1)


def _default_queries():
    """Four queries, not six. Each costs API credits and the free monthly
    allowance is modest — two extra phrasings mostly returned the same roles."""
    a, b, nxt = _season_terms()
    return [
        f"software engineering internship Auckland New Zealand summer {a} apply",
        f"software developer internship Auckland student {b}",
        f"computer science internship Auckland summer {nxt} applications open",
        f"IT data embedded engineering internship Auckland student {nxt}",
    ]


DEFAULT_QUERIES = _default_queries()

# ---------------------------------------------------------------------------
# Eligibility — the bit that saves you the scrolling
# ---------------------------------------------------------------------------

# Wording that means "not you (yet)". Ordered roughly by how decisive it is.
_EXCLUDE = [
    ("penultimate", r"\bpenultimate\b"),
    ("final year only", r"\b(final|last)[\s-]?year\s+(student|students|only|"
                       r"undergraduate|candidates)?\b"),
    ("3rd/4th year", r"\b(third|fourth|3rd|4th)[\s-]?year\b"),
    ("graduating soon", r"\bgraduat\w*\s+(in|by)\s+20(2[6-8])\b|"
                        r"\bdue to graduate\b|\bcompleting your degree\b"),
    ("must have degree", r"\b(completed|finished|hold)\s+(a\s+)?(bachelor|degree)\b|"
                         r"\bdegree required\b"),
    ("senior role", r"\b(senior|lead|principal|staff)\s+(software\s+)?"
                    r"(engineer|developer)\b"),
    ("postgrad only", r"\b(phd|doctoral|masters?[\s-]?(student|degree|level)|"
                      r"postgraduate|post[\s-]?doc\w*)\b"),
    ("years experience", r"\b([3-9]|1\d)\+?\s*years?\s+(of\s+)?"
                         r"(commercial\s+|professional\s+|industry\s+)?experience\b"),
]

# ---------------------------------------------------------------------------
# Closed and stale postings
#
# Boards keep old adverts up for years. An EROAD listing from the 2025/2026
# season was still being surfaced in July 2026, marked "No longer accepting
# applications" — a dead link wastes more of your time than an ineligible one,
# because you only find out after clicking through.
# ---------------------------------------------------------------------------

_CLOSED = re.compile(
    r"\bno longer accepting applications?\b|\bapplications? (?:are )?closed\b|"
    r"\bthis (?:job|position|role|vacancy) is no longer\b|"
    r"\bposition (?:has been )?filled\b|\bnot accepting applications\b|"
    r"\bexpired\b|\bclosed for applications\b|\brecruitment (?:has )?closed\b|"
    r"\bapplications have closed\b|"
    # aggregators archive old rounds under headings like "Past Internships"
    r"\bpast (?:graduate|intern|job|opportunit|program)\w*\b|"
    r"\bprevious(?:ly)? (?:advertised|listed)\b|\bnow closed\b", re.I)

# "1 year ago", "11 months ago" — boards print this next to the post date.
_POSTED_AGE = re.compile(r"\b(\d+)\+?\s*(day|week|month|year)s?\s+ago\b", re.I)

# "2025 - 2026", "2025/26", "2026/2027" — an NZ summer season.
_SEASON = re.compile(r"\b(20\d{2})\s*[-/–]\s*(20\d{2}|\d{2})\b")

_MAX_AGE_DAYS = 150          # ~5 months; summer campaigns open around then


def posting_age_days(text: str):
    """Days since posting, if the page says. None when it doesn't."""
    best = None
    for n, unit in _POSTED_AGE.findall(text or ""):
        try:
            n = int(n)
        except ValueError:
            continue
        days = n * {"day": 1, "week": 7, "month": 30, "year": 365}[unit.lower()]
        best = days if best is None else min(best, days)
    return best


def season_is_past(text: str) -> bool:
    """Does this advertise a summer season that has already finished?

    NZ summer internships run roughly November to February, so the 2025/2026
    season is over by March 2026 — an advert naming it in July 2026 is stale.
    """
    import datetime
    today = datetime.date.today()
    for start, end in _SEASON.findall(text or ""):
        try:
            start = int(start)
            end = int(end) if len(end) == 4 else int(str(start)[:2] + end)
        except ValueError:
            continue
        if end < today.year:
            return True
        if end == today.year and today.month > 3:
            return True
    return False


_VERIFY_UA = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                  "(KHTML, like Gecko) Chrome/122 Safari/537.36",
    "Accept-Language": "en-NZ,en;q=0.9",
}


def verify_open(url: str) -> dict:
    """Fetch the posting itself and see whether it's actually live.

    This is the only reliable check. Tavily's extraction routinely omits the
    status line, so closed roles were being shown as open — fetching the page
    directly finds "No longer accepting applications" on postings that looked
    fine, and a dead advert 404s outright.

    Returns {"open": bool, "reason": str, "checked": bool}; checked is False
    when the fetch failed, in which case the posting is left alone rather than
    hidden on a network hiccup.
    """
    if not url:
        return {"open": True, "reason": "", "checked": False}
    try:
        import requests
        r = requests.get(url, headers=_VERIFY_UA, timeout=12,
                         allow_redirects=True)
    except Exception:
        return {"open": True, "reason": "", "checked": False}

    if r.status_code in (404, 410):
        return {"open": False, "reason": "posting removed", "checked": True}
    if r.status_code >= 400:
        return {"open": True, "reason": "", "checked": False}

    text = re.sub(r"<[^>]+>", " ", r.text)
    text = re.sub(r"\s+", " ", text)
    head = _posting_only(text)[:6000]
    if _CLOSED.search(head):
        return {"open": False, "reason": "closed", "checked": True}
    age = posting_age_days(head)
    if age is not None and age > _MAX_AGE_DAYS:
        return {"open": False, "reason": f"posted {round(age / 30)} months ago",
                "checked": True}
    return {"open": True, "reason": "", "checked": True}


def _verify_all(rows):
    """Check the shortlist in parallel — one fetch each, ~1-2s overall."""
    out, lock = {}, threading.Lock()

    def go(row):
        v = verify_open(row["url"])
        with lock:
            out[row["url"]] = v

    threads = [threading.Thread(target=go, args=(r,), daemon=True) for r in rows]
    for t in threads:
        t.start()
    for t in threads:
        t.join(timeout=15)
    return out


def check_freshness(text: str, title: str = "") -> dict:
    """Is this still open and current? {open, reason}"""
    blob = f"{title}\n{text or ''}"
    if _CLOSED.search(blob):
        return {"open": False, "reason": "closed"}
    age = posting_age_days(blob)
    if age is not None and age > _MAX_AGE_DAYS:
        months = round(age / 30)
        return {"open": False, "reason": f"posted {months} months ago"}
    if season_is_past(f"{title} {(text or '')[:600]}"):
        return {"open": False, "reason": "past season"}
    return {"open": True, "reason": ""}


# Wording that explicitly welcomes earlier years — overrides a weak exclusion.
_INCLUSIVE = re.compile(
    r"\b(all years|any year|first[\s-]?year|second[\s-]?year|1st[\s-]?year|"
    r"2nd[\s-]?year|early[\s-]?career|no experience (required|necessary)|"
    r"students at any stage|open to all undergraduate)\b", re.I)

# Must look like software/computing, not any old internship.
# Trailing \w* on the stems that take plurals — \bmechatronic\b does not match
# "Mechatronics", which silently dropped exactly the roles you asked to include.
_RELEVANT = re.compile(
    r"\b(software|developer|develop\w*|programming|programmer|comput\w*|"
    r"data (?:science|engineer\w*|analyst)|machine learning|ml|ai|"
    r"web|full[\s-]?stack|back[\s-]?end|front[\s-]?end|devops|cloud|"
    r"embedded|firmware|mechatronic\w*|robotic\w*|electrical engineer\w*|"
    r"electronic\w*|cyber\w*|security engineer|qa|test engineer|"
    r"it|information tech\w*|tech\w*)\b",
    re.I)

_INTERNSHIP = re.compile(
    r"\b(intern|internship|placement|summer student|work experience|"
    r"industrial experience|co[\s-]?op|cadet\w*|trainee|graduate programme|"
    r"summer of tech)\b", re.I)

# Auckland specifically, or genuinely remote.
_AUCKLAND = re.compile(
    r"\b(auckland|tamaki makaurau|tāmaki makaurau|albany|takapuna|"
    r"manukau|penrose|newmarket|north shore|east tamaki|henderson|"
    r"mount wellington|parnell|wynyard quarter)\b", re.I)

_REMOTE = re.compile(
    r"\b(remote|work from home|wfh|anywhere in (?:nz|new zealand)|"
    r"fully distributed|location independent)\b", re.I)

# Somewhere else. Australia is the big one — the feeds are full of Sydney and
# Melbourne roles — but other NZ cities matter too if you're not moving.
_ELSEWHERE = re.compile(
    r"\b(sydney|melbourne|brisbane|perth|adelaide|canberra|hobart|darwin|"
    r"gold coast|australia|australian|nsw|victoria, au|queensland|"
    r"wellington|christchurch|hamilton|dunedin|tauranga|palmerston north|"
    r"napier|nelson|queenstown|rotorua|new plymouth|invercargill|whangarei|"
    r"singapore|london|new york|san francisco|bangalore|india|manila)\b", re.I)


def location_ok(text: str) -> bool:
    """Auckland, or remote — and not somewhere else wearing an NZ label.

    "New Zealand" alone used to pass, which let Wellington and Hamilton roles
    through, and the Discord feed carries as many Sydney posts as Auckland
    ones. Auckland (or explicit remote) now has to be named, and it has to
    out-rank any other place mentioned.
    """
    blob = text or ""
    here = bool(_AUCKLAND.search(blob))
    remote = bool(_REMOTE.search(blob))
    other = bool(_ELSEWHERE.search(blob))
    if here:
        return True                 # names Auckland: good even if it lists others
    if remote and not other:
        return True                 # remote, with nowhere else claimed
    return False


# Job pages carry a sidebar of OTHER roles ("Similar jobs", "People also
# viewed"). Reading past this point hid an Apple *intern* role as a "senior
# role", because a senior listing appeared further down the same page.
_BOILERPLATE = re.compile(
    r"\b(similar jobs|people also viewed|more jobs|related jobs|"
    r"recommended for you|jobs you may be interested|explore collaborative|"
    r"sign in to|create job alert|referrals increase)\b", re.I)

# These describe the ROLE, so they're only trustworthy in the title. Body text
# routinely says "you'll be mentored by senior engineers".
_TITLE_ONLY = {"senior role", "postgrad only"}


def _posting_only(text: str) -> str:
    """The advert itself, with the page's other-jobs furniture cut off."""
    m = _BOILERPLATE.search(text or "")
    return (text or "")[:m.start()] if m else (text or "")


def check_eligibility(text: str, title: str = "") -> dict:
    """Would this posting take a 2nd-year? Returns {eligible, reason}."""
    body = _posting_only(text or "")
    head = title or ""
    inclusive = bool(_INCLUSIVE.search(body))
    for label, pattern in _EXCLUDE:
        target = head if label in _TITLE_ONLY else body
        if re.search(pattern, target, re.I):
            # An explicit welcome for early years beats a soft signal, but
            # never beats "penultimate" or a hard year requirement.
            if inclusive and label in ("must have degree", "graduating soon",
                                       "years experience"):
                continue
            return {"eligible": False, "reason": label}
    return {"eligible": True,
            "reason": "welcomes earlier years" if inclusive else "no year restriction found"}


def looks_relevant(text: str) -> bool:
    return bool(_RELEVANT.search(text or ""))


def looks_like_internship(text: str) -> bool:
    return bool(_INTERNSHIP.search(text or ""))




# ---------------------------------------------------------------------------
# Tidying a search result into something that reads like a listing
# ---------------------------------------------------------------------------

# "Serko hiring Intern Software Engineer - Summer 2025/2026 in Auckland"
_HIRING = re.compile(r"^(?P<company>.+?)\s+hiring\s+(?P<title>.+?)"
                     r"(?:\s+in\s+(?P<where>.+))?$", re.I)
_LISTING_PAGE = re.compile(
    r"^\d[\d,]*\+?\s|jobs? in |job vacancies|search \d|"
    r"\bjobs,? employment\b|\bjob search\b", re.I)

# A company's careers landing page — not a specific role, but still worth
# knowing about for a company you're targeting. Shown below real postings.
_CAREERS_PAGE = re.compile(
    r"\bcareers?\b|\bcurrent vacanc\w+\b|\bjob opportunit\w+\b|"
    r"\bwork (?:with|for) us\b|\bjoin (?:us|our team)\b|\bearly careers?\b|"
    r"\bopportunities\b|\bvacancies\b", re.I)

# Content farms and listicles that rank well and help not at all.
_CONTENT_FARM = re.compile(
    r"\bnucamp\b|\btop \d+\b|\bbest \d+\b|\b\d+ best\b|\bultimate guide\b|"
    r"\bhow to (?:get|land|find)\b|\bblog\b|\bwhat is\b|\bcourse\b|\bbootcamp\b",
    re.I)

# Social and forum results are never the advert itself.
_BLOCK_DOMAINS = ("facebook.com", "twitter.com", "x.com", "reddit.com",
                  "youtube.com", "instagram.com", "tiktok.com", "quora.com",
                  "glassdoor.", "medium.com", "wikipedia.org")

# A URL that ends at a section index rather than a specific advert.
_INDEX_URL = re.compile(
    r"/(careers?|jobs?|internships?|opportunities|vacancies|"
    r"work-with-us|join-us|early-careers?)/?$", re.I)


def _parse_title(title: str, url: str) -> dict:
    t = (title or "").strip()
    # DuckDuckGo prefixes a breadcrumb — "nz.linkedin.com › jobs › view…" —
    # which was ending up as the company name. Keep the last segment; a regex
    # that stripped the prefix ate the whole title and silently fell back.
    if "›" in t:
        tail = t.split("›")[-1].strip()
        # "viewSoftware Engineering Intern" — the crumb runs into the title
        tail = re.sub(r"^(?:view|jobs?|careers?|search)(?=[A-Z])", "", tail)
        tail = re.sub(r"^\d{4,}(?=[A-Za-z])", "", tail)   # seek: "12345Intern…"
        if len(tail) > 4:
            t = tail
    t = re.sub(r"\s*[-|]\s*(LinkedIn|Seek|Indeed|Jobs?|Careers?)\s*$", "", t,
               flags=re.I).strip() or t
    company, role, where = "", t, ""
    m = _HIRING.match(t)
    if m:
        company = m.group("company").strip()
        role = m.group("title").strip()
        where = (m.group("where") or "").strip()
    else:
        # "Role at Company" / "Role - Company"
        m2 = re.match(r"^(?P<title>.+?)\s+(?:at|@|\|)\s+(?P<company>[^|]+)$", t)
        if m2:
            role, company = m2.group("title").strip(), m2.group("company").strip()
    host = urllib.parse.urlparse(url).netloc.replace("www.", "")
    return {"company": company, "role": role, "where": where, "source": host}


def _is_individual_posting(url: str, title: str) -> bool:
    """A specific advert, rather than a page that lists adverts.

    Careers landing pages were being shown and were no use — clicking one just
    lands you on a site full of jobs, which is the position you started from.
    """
    if _LISTING_PAGE.search(title or ""):
        return False
    u = (url or "").lower()
    if _INDEX_URL.search(u):
        return False
    # a specific posting nearly always carries an id in the path or query
    if re.search(r"/jobs?/view/|/job/\d|/jobs?/\d|/vacanc\w*/\d|jobid=|"
                 r"/positions?/\d|/opening/\d|[?&](?:gh_jid|lever|jid|id)=", u):
        return True
    if re.search(r"seek\.co\.nz/job/\d+", u):
        return True
    # applicant-tracking hosts put the role in the path (workable, lever,
    # greenhouse, bamboo); accept when the title names an actual role
    if re.search(r"(workable|lever\.co|greenhouse\.io|bamboohr|smartrecruiters|"
                 r"recruitee|teamtailor|jobvite|myworkday)", u):
        return bool(_INTERNSHIP.search(title or ""))
    return False


# ---------------------------------------------------------------------------
# Search
# ---------------------------------------------------------------------------

_cache = {"at": 0.0, "items": [], "skipped": [], "coverage": {}}
_lock = threading.Lock()
_searching = [False]

# ---------------------------------------------------------------------------
# Dismissed listings
#
# Kept on disk so "seen" survives a restart — a job board reposts the same
# roles for weeks, and re-reading them is the thing this feature exists to
# avoid. Keyed by URL, which is stable per posting.
# ---------------------------------------------------------------------------

SEEN_PATH = os.path.join(BASE, "config", "jobs_seen.json")
_seen_lock = threading.Lock()
_seen_cache = {"at": 0.0, "data": None}


def _load_seen() -> dict:
    with _seen_lock:
        if _seen_cache["data"] is not None:
            return _seen_cache["data"]
        try:
            import json
            with open(SEEN_PATH, encoding="utf-8") as f:
                data = json.load(f)
            if not isinstance(data, dict):
                data = {}
        except Exception:
            data = {}
        _seen_cache["data"] = data
        return data


def _save_seen(data: dict):
    import json
    tmp = SEEN_PATH + ".tmp"
    try:
        os.makedirs(os.path.dirname(SEEN_PATH), exist_ok=True)
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2)
            f.write("\n")
        os.replace(tmp, SEEN_PATH)
    except Exception:
        pass


def mark_seen(url: str, role: str = "") -> dict:
    """Hide one listing from future results."""
    url = (url or "").split("?")[0]
    if not url:
        return {"ok": False, "reply": "No listing given"}
    data = _load_seen()
    data[url] = {"role": role[:120], "at": time.strftime("%Y-%m-%d %H:%M")}
    with _seen_lock:
        _seen_cache["data"] = data
    _save_seen(data)
    # drop it from the cached page too, so the row goes immediately
    with _lock:
        _cache["items"] = [i for i in _cache["items"] if i.get("url") != url]
    return {"ok": True, "seen": len(data),
            "reply": f"Hidden — {role[:60]}" if role else "Hidden"}


def unmark_seen(url: str = None) -> dict:
    """Bring one back, or all of them if no url is given."""
    data = _load_seen()
    if url:
        data.pop((url or "").split("?")[0], None)
    else:
        data = {}
    with _seen_lock:
        _seen_cache["data"] = data
    _save_seen(data)
    with _lock:
        _cache["at"] = 0.0          # force a fresh search so they reappear
    return {"ok": True, "seen": len(data),
            "reply": "Showing all listings again" if not url else "Restored"}


def seen_count() -> int:
    return len(_load_seen())


# ---------------------------------------------------------------------------
# Remembered eligibility verdicts
#
# Tavily doesn't always return the same amount of a page, so the same posting
# can look ineligible on one search and fine on the next — Fisher & Paykel's
# 3rd/4th-year role was hidden once and then came back top-ranked. Once a hard
# year requirement has been read from a posting, remember it: a role does not
# stop requiring 3rd year because the fetch was shorter this time.
# ---------------------------------------------------------------------------

VERDICT_PATH = os.path.join(BASE, "config", "jobs_verdicts.json")
_verdict_cache = {"data": None}

# Only durable facts about the posting are worth remembering. "no year
# restriction found" may simply mean we didn't get that far down the page.
_STICKY = {"penultimate", "final year only", "3rd/4th year", "graduating soon",
           "must have degree", "postgrad only", "senior role",
           "closed", "past season", "posting removed"}


def _is_sticky(reason: str) -> bool:
    # "posted 14 months ago" only gets older, so remember those too
    return reason in _STICKY or reason.startswith("posted ")


def _load_verdicts() -> dict:
    if _verdict_cache["data"] is not None:
        return _verdict_cache["data"]
    try:
        import json
        with open(VERDICT_PATH, encoding="utf-8") as f:
            data = json.load(f)
        if not isinstance(data, dict):
            data = {}
    except Exception:
        data = {}
    _verdict_cache["data"] = data
    return data


def _remember_verdict(url: str, reason: str):
    if not _is_sticky(reason):
        return
    data = _load_verdicts()
    if data.get(url) == reason:
        return
    data[url] = reason
    _verdict_cache["data"] = data
    try:
        import json
        tmp = VERDICT_PATH + ".tmp"
        os.makedirs(os.path.dirname(VERDICT_PATH), exist_ok=True)
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2)
            f.write("\n")
        os.replace(tmp, VERDICT_PATH)
    except Exception:
        pass


def _sites():
    return config.get("jobs", "sites", default=DEFAULT_SITES) or DEFAULT_SITES


def _queries():
    # recomputed each call so the season rolls over without editing settings
    return config.get("jobs", "queries", default=None) or _default_queries()


# Set when the search API refuses us, so the UI can say why instead of
# reporting "nothing found" — a silent empty result is indistinguishable from
# "there are no jobs", which is exactly the wrong thing to tell someone.
_last_error = {"msg": ""}


def last_error() -> str:
    return _last_error["msg"]


def _tavily_search(query: str, sites, time_range: str = "year"):
    from skills import web
    key = web._tavily_key()
    if not key:
        _last_error["msg"] = "No Tavily API key set (web.tavily_api_key)."
        return []
    try:
        from tavily import TavilyClient
        r = TavilyClient(api_key=key).search(
            query=query, search_depth="advanced", max_results=20,
            country="new zealand", include_domains=list(sites),
            # bias to recent pages: boards keep years of dead adverts indexed
            time_range=time_range,
            # The year requirement lives in the posting body, not the snippet.
            # Without the full page the eligibility filter had nothing to read
            # and hid nothing at all — which is the whole point of this.
            include_raw_content=True)
        return r.get("results") or []
    except Exception as e:
        name = type(e).__name__
        if "UsageLimit" in name or "quota" in str(e).lower():
            _last_error["msg"] = ("Tavily monthly credits are used up — "
                                  "using keyless search instead.")
        elif "Invalid" in name or "401" in str(e):
            _last_error["msg"] = "Tavily rejected the API key."
        else:
            _last_error["msg"] = f"Search failed — {name}"
        return []


def _ddgs_jobs(query: str, sites):
    """Keyless fallback. Not as good as Tavily — no page content, so the
    eligibility filter has less to read — but it keeps working when the
    Tavily allowance runs out, instead of reporting 'no jobs found'."""
    try:
        try:
            from ddgs import DDGS
        except Exception:
            from duckduckgo_search import DDGS
        out, lock = [], threading.Lock()

        def one(site):
            try:
                with DDGS() as ddg:
                    hits = list(ddg.text(f"{query} site:{site}", max_results=8))
            except Exception:
                return
            with lock:
                for h in hits:
                    out.append({"title": h.get("title") or "",
                                "url": h.get("href") or h.get("link") or "",
                                "content": h.get("body") or "",
                                "raw_content": "", "score": 0.5})

        # in parallel: run one site at a time this took over a minute
        th = [threading.Thread(target=one, args=(s,), daemon=True)
              for s in list(sites)[:4]]
        for t in th:
            t.start()
        for t in th:
            t.join(timeout=20)
        return out
    except Exception:
        return []


def _search_all(queries, sites):
    """Run the queries at once. Sequentially this took 73s.

    Each query runs twice: once over the last month and once over the year.
    Search engines index dead adverts for years — an audit of what came back
    found all but one posting already closed — so the recent pass is what
    actually surfaces roles you can still apply to, while the wider pass keeps
    coverage for boards that don't date their pages.
    """
    out, lock = [], threading.Lock()
    _last_error["msg"] = ""

    def go(q, window):
        hits = _tavily_search(q, sites, time_range=window)
        with lock:
            out.extend(hits)

    # One window per query, alternating, rather than both for every query:
    # running both doubled the credit cost for a modest gain, and the free
    # monthly allowance is small enough that it ran out mid-development.
    pairs = [(q, "month" if i % 2 == 0 else "year")
             for i, q in enumerate(queries)]
    threads = [threading.Thread(target=go, args=(q, w), daemon=True)
               for q, w in pairs]
    for t in threads:
        t.start()
    for t in threads:
        t.join(timeout=45)

    # Tavily unavailable (no key, or the allowance is gone): fall back to the
    # keyless search rather than pretending there are no jobs.
    if not out and _last_error["msg"]:
        for q in queries[:3]:
            out.extend(_ddgs_jobs(q, sites))
    return out


def known_companies() -> list:
    """Companies you've noted as having hired software people before.

    Point `jobs.companies_file` at that spreadsheet (first column = name, or a
    column headed Company). They get their own targeted search and rank above
    everything else, since a company that's hired before is worth more than a
    generic board hit.
    """
    path = os.path.expandvars(config.get("jobs", "companies_file", default="") or "")
    if not path or not os.path.isfile(path):
        return []
    out, seen = [], set()
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        for sheet in wb.sheetnames:          # e.g. "Software Internships"
            rows = list(wb[sheet].iter_rows(values_only=True))   # + "All Companies"
            if not rows:
                continue
            header = [str(c or "").strip().lower() for c in rows[0]]
            # "Company Name", "Company", "Employer" — match loosely
            col = next((i for i, h in enumerate(header)
                        if "company" in h or "employer" in h or "organisation" in h), 0)
            for r in rows[1:]:
                if col >= len(r) or not r[col]:
                    continue
                name = str(r[col]).strip()
                if name and len(name) < 60 and name.lower() not in seen:
                    seen.add(name.lower())
                    out.append(name)
        wb.close()
    except Exception:
        return []
    return out[:80]


# ---------------------------------------------------------------------------
# Company career pages
#
# Job boards are only half the picture: plenty of the companies in your list
# advertise on their own careers site and nowhere else. Board search can't see
# those at all, because include_domains restricts results to the boards.
#
# Checking 80 companies on every refresh would burn the Tavily allowance, so
# each company is cached for a long TTL and only the stalest handful are
# re-checked per run. Over a few refreshes the whole list gets covered, and
# coverage is reported so it's clear how far through it is.
# ---------------------------------------------------------------------------

CAREERS_PATH = os.path.join(BASE, "config", "jobs_companies.json")
_careers_cache = {"data": None}


def _load_careers() -> dict:
    if _careers_cache["data"] is not None:
        return _careers_cache["data"]
    try:
        import json
        with open(CAREERS_PATH, encoding="utf-8") as f:
            data = json.load(f)
        if not isinstance(data, dict):
            data = {}
    except Exception:
        data = {}
    _careers_cache["data"] = data
    return data


def _save_careers(data: dict):
    try:
        import json
        tmp = CAREERS_PATH + ".tmp"
        os.makedirs(os.path.dirname(CAREERS_PATH), exist_ok=True)
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=1)
            f.write("\n")
        os.replace(tmp, CAREERS_PATH)
    except Exception:
        pass


def _career_search(company: str):
    """Look at the company's own site, not just the boards."""
    from skills import web
    key = web._tavily_key()
    if not key:
        return []
    try:
        from tavily import TavilyClient
        r = TavilyClient(api_key=key).search(
            query=f"{company} careers internship software engineering "
                  f"New Zealand student 2026 apply",
            # basic depth here: this runs across many companies, and the
            # advanced tier costs double per search
            search_depth="basic", max_results=5, country="new zealand",
            include_raw_content=True)
        return r.get("results") or []
    except Exception:
        return []


def companies_due(known, limit):
    """The stalest companies, so every one gets covered in rotation."""
    ttl = float(config.get("jobs", "company_ttl_hours", default=48) or 48) * 3600
    seen_at = _load_careers()
    now = time.time()
    stale = [(seen_at.get(c, {}).get("at", 0.0), c) for c in known]
    stale = [(at, c) for at, c in stale if now - at > ttl]
    stale.sort()
    return [c for _, c in stale[:max(0, limit)]]


def coverage(known) -> dict:
    ttl = float(config.get("jobs", "company_ttl_hours", default=48) or 48) * 3600
    data = _load_careers()
    now = time.time()
    done = sum(1 for c in known if now - data.get(c, {}).get("at", 0.0) <= ttl)
    return {"checked": done, "total": len(known), "pending": len(known) - done}


def _applied_companies():
    """Don't re-suggest somewhere you've already applied."""
    try:
        from skills import excel_sync
        return {c.strip().lower() for c in excel_sync.companies() if c}
    except Exception:
        return set()


def search(force: bool = False, max_results: int = 25) -> dict:
    """Find internships you're actually eligible for."""
    ttl = float(config.get("jobs", "cache_hours", default=6) or 6) * 3600
    with _lock:
        if not force and _cache["items"] and (time.time() - _cache["at"]) < ttl:
            return _result(_cache["items"], _cache["skipped"], cached=True)

    sites = _sites()
    seen, items, skipped = {}, [], []
    applied = _applied_companies()
    dismissed = set(_load_seen())
    remembered = _load_verdicts()

    known = known_companies()
    known_lc = {k.lower() for k in known}

    # Pass 1 — the job boards.
    hits = _search_all(_queries(), sites)

    # Pass 2 — the companies' own careers pages, a rotating batch so the whole
    # list is covered over successive refreshes without burning the allowance.
    batch = int(config.get("jobs", "company_batch", default=10) or 10)
    due = companies_due(known, batch)
    if due:
        careers = _load_careers()
        results, lock = {}, threading.Lock()

        def go(name):
            found = _career_search(name)
            with lock:
                results[name] = found

        threads = [threading.Thread(target=go, args=(c,), daemon=True) for c in due]
        for t in threads:
            t.start()
        for t in threads:
            t.join(timeout=40)

        for name, found in results.items():
            careers[name] = {"at": time.time(), "n": len(found)}
            for f in found:
                f["_company_hint"] = name
                hits.append(f)
        _careers_cache["data"] = careers
        _save_careers(careers)

    for hit in hits:
            url = (hit.get("url") or "").split("?")[0]
            title = hit.get("title") or ""
            if not url or url in seen:
                continue
            seen[url] = True
            if url in dismissed:
                continue          # you've already looked at this one
            # raw_content is the full posting; that's where "penultimate" and
            # "final year" actually appear.
            raw = (hit.get("raw_content") or "")[:12000]
            body = f"{title}\n{hit.get('content') or ''}\n{raw}"

            # Career-page results are kept even when the URL doesn't look like
            # a board posting — "/careers/", "/join-us" and the like are
            # exactly what a board search can't reach.
            # Only actual adverts. A careers index or a board's search page is
            # just a link to more searching, which is what you already have.
            if not _is_individual_posting(url, title):
                continue
            if _INDEX_URL.search(url) or _LISTING_PAGE.search(title or ""):
                continue
            # listicles, course ads and social posts rank well and help not at all
            if _CONTENT_FARM.search(title or "") or _CONTENT_FARM.search(url):
                continue
            if any(d in url.lower() for d in _BLOCK_DOMAINS):
                continue
            if not looks_like_internship(body) or not looks_relevant(body):
                continue
            if not location_ok(body):
                continue

            parsed = _parse_title(title, url)
            # the careers pass knows the company even when the page title
            # doesn't spell it out
            if hit.get("_company_hint") and not parsed["company"]:
                parsed["company"] = hit["_company_hint"]
            elig = check_eligibility(body, title=f"{title} {parsed['role']}")
            # a requirement read once still applies, even if this fetch was
            # shorter and didn't include it
            # Closed or out-of-season before anything else: a dead link is
            # worse than an ineligible one, because you only find out after
            # clicking through.
            fresh = check_freshness(body, title)
            if not fresh["open"]:
                elig = {"eligible": False, "reason": fresh["reason"]}
            prior = remembered.get(url)
            if elig["eligible"] and prior:
                elig = {"eligible": False, "reason": prior}
            elif not elig["eligible"]:
                _remember_verdict(url, elig["reason"])
            row = {
                "role": parsed["role"][:110],
                "company": parsed["company"][:60],
                "where": parsed["where"][:60],
                "source": parsed["source"],
                "url": url,
                "score": round(float(hit.get("score") or 0), 3),
                "reason": elig["reason"],
                "already_applied": parsed["company"].strip().lower() in applied,
                "known_hirer": parsed["company"].strip().lower() in known_lc,
                # a landing page rather than a specific role: still useful for
                # a company you're targeting, but it shouldn't outrank a job
                # a section index, or a title that names no specific role
                "careers_page": bool(
                    _INDEX_URL.search(url)
                    or _LISTING_PAGE.search(parsed["role"])
                    or (_CAREERS_PAGE.search(parsed["role"])
                        and not re.search(r"\b(intern|engineer|developer|"
                                          r"analyst|graduate)\b",
                                          parsed["role"], re.I))),
            }
            if row["known_hirer"]:
                row["score"] = round(row["score"] + 0.15, 3)   # rank these up
            (items if elig["eligible"] else skipped).append(row)

    # Boards list the same role more than once (reposts, multiple locations).
    # Keep the best-scoring copy of each company+role.
    best = {}
    for row in items:
        key = (row["company"].lower().strip(),
               re.sub(r"[^a-z0-9]", "", row["role"].lower())[:40])
        if key not in best or row["score"] > best[key]["score"]:
            best[key] = row
    items = sorted(best.values(),
                   key=lambda r: (r["already_applied"], -r["score"]))[:max_results]

    # Confirm each survivor is genuinely still open by fetching it. Done last,
    # on the shortlist only, so it's ~15 requests rather than hundreds.
    if config.get("jobs", "verify_open", default=True):
        verdicts = _verify_all(items)
        live = []
        for row in items:
            v = verdicts.get(row["url"], {})
            if v.get("checked") and not v.get("open"):
                row["reason"] = v["reason"]
                skipped.append(row)
                _remember_verdict(row["url"], v["reason"])
            else:
                row["verified"] = bool(v.get("checked"))
                live.append(row)
        items = live

    # The Discord feed is a person posting roles the day they open, so it goes
    # in alongside the crawled results rather than in a separate list.
    try:
        from skills import discord_jobs
        if discord_jobs.is_configured():
            d = discord_jobs.recent()
            have = {i["url"] for i in items}
            items.extend(x for x in d.get("items", []) if x["url"] not in have)
            skipped.extend(d.get("skipped", []))
            items.sort(key=lambda r: (r["already_applied"], -r["score"]))
    except Exception:
        pass

    cov = coverage(known)
    with _lock:
        _cache["at"] = time.time()
        _cache["items"] = items
        _cache["skipped"] = skipped
        _cache["coverage"] = cov
    return _result(items, skipped, cached=False, cov=cov)


def _result(items, skipped, cached, cov=None):
    fresh = [i for i in items if not i["already_applied"] and not i.get("careers_page")]
    n_seen = seen_count()
    err = last_error()
    if not items:
        if err:
            # never report "no jobs" when the search itself didn't run
            reply = err
        else:
            reply = ("Nothing new — everything matching is either already seen "
                     "or not open to your year." if n_seen else
                     "No internships matched right now.")
    else:
        pages = sum(1 for i in items if i.get("careers_page"))
        top = fresh[0] if fresh else items[0]
        who = top["company"] or top["source"]
        reply = f"{len(fresh)} internship(s) you're eligible for · top: {top['role']}"
        if who:
            reply += f" at {who}"
        if skipped:
            reply += f" · {len(skipped)} hidden (penultimate/final-year)"
        if pages:
            reply += f" · {pages} careers page(s) to check yourself"
        if n_seen:
            reply += f" · {n_seen} dismissed"
    cov = cov or _cache.get("coverage") or {}
    if cov.get("total"):
        reply += (f" · career pages: {cov['checked']}/{cov['total']} companies"
                  + (f", {cov['pending']} still to check" if cov.get("pending") else ""))
    else:
        if err:
            reply += f" · note: {err}"
    return {"ok": True, "intent": "jobs", "items": items, "skipped": skipped,
            "cached": cached, "seen": n_seen, "coverage": cov,
            "error": err, "reply": reply}


def warm():
    """Refresh in the background so the Jobs tab is populated when opened."""
    if _searching[0]:
        return
    _searching[0] = True

    def go():
        try:
            search()
        except Exception:
            pass
        finally:
            _searching[0] = False

    threading.Thread(target=go, daemon=True).start()


def is_ready() -> bool:
    return bool(_cache["items"])


def cached() -> dict:
    return _result(_cache["items"], _cache["skipped"], cached=True)
